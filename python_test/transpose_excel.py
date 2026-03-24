"""
Script to transpose rows to columns in an Excel file
"""

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def transpose_excel_file(input_file, output_file, sheet_name=None):
    """
    Transpose data from rows to columns in an Excel file
    
    Args:
        input_file (str): Path to the input Excel file
        output_file (str): Path to save the transposed Excel file
        sheet_name (str): Name of the sheet to transpose (defaults to first sheet)
    """
    
    # Load the input workbook
    wb = load_workbook(input_file)
    
    # Get the sheet (default to first sheet if not specified)
    if sheet_name is None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    
    # Read all data from the sheet
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    # Create a new workbook for transposed data
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Transposed"
    
    # Get the dimensions
    num_rows = len(data)
    num_cols = max(len(row) for row in data) if data else 0
    
    # Transpose the data
    for i in range(num_cols):
        for j in range(num_rows):
            # Get value from original position (row j, col i)
            # Write to transposed position (row i, col j)
            if j < len(data) and i < len(data[j]):
                value = data[j][i]
                new_ws.cell(row=i+1, column=j+1, value=value)
    
    # Save the new workbook
    new_wb.save(output_file)
    
    print(f"✓ Transposed data saved to: {output_file}")
    print(f"  Original dimensions: {num_rows} rows × {num_cols} columns")
    print(f"  Transposed dimensions: {num_cols} rows × {num_rows} columns")


if __name__ == "__main__":
    # Example usage
    input_file = "sample_data.xlsx"
    output_file = "transposed_data.xlsx"
    
    try:
        transpose_excel_file(input_file, output_file)
        print("\nTransposition completed successfully!")
    except FileNotFoundError:
        print(f"Error: {input_file} not found. Please ensure the file exists.")
    except Exception as e:
        print(f"Error: {str(e)}")
